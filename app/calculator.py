# -*- coding: utf-8 -*-
"""
Logika obliczeniowa kalkulatora okleinowania
"""

from typing import List, Dict, Tuple
from dataclasses import dataclass, field


@dataclass
class Profil:
    """Reprezentacja profilu okleinowania"""
    nazwa: str
    dlg_szt: float  # długość w mm
    ilosc: int      # ilość sztuk
    color: str      # np. "RAL 4R7021 MAT*STANDARD"
    wybrany: bool = True
    
    def get_mb(self) -> float:
        """Oblicz metraż bieżący"""
        return (self.dlg_szt / 1000) * self.ilosc
    
    def calculate_price(self, kwota_okleinowania: float, dwustronne: bool = False, 
                       wspolczynnik: float = 1.0) -> float:
        """Oblicz cenę profilu"""
        if not self.wybrany:
            return 0.0
        
        mb = self.get_mb()
        cena = mb * kwota_okleinowania
        
        if dwustronne:
            cena *= wspolczynnik
        
        # Dodaj opłatę za profile poniżej 300 zł
        if cena < 300:
            cena += 90
        
        return cena


class OkleinowanieCalculator:
    """Kalkulator okleinowania profili"""
    
    def __init__(self):
        self.profile: List[Profil] = []
    
    def load_from_excel(self, filepath: str, sheet_name: str = None) -> bool:
        """Wczytaj profile z pliku Excel
        
        Args:
            filepath: ścieżka do pliku .xlsx
            sheet_name: nazwa arkusza (jeśli None, użyj pierwszego)
        
        Returns:
            True jeśli udało się wczytać, False w przeciwnym razie
        """
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("Biblioteka 'openpyxl' nie jest zainstalowana. Zainstaluj ją: pip install openpyxl")
        
        try:
            wb = load_workbook(filepath)
            
            if sheet_name is None:
                ws = wb.active
            else:
                ws = wb[sheet_name]
            
            self.profile = []
            
            # Przeszukaj pierwszy wiersz w poszukiwaniu nagłówków
            headers = {}
            for col_idx, cell in enumerate(ws[1], 1):
                if cell.value:
                    headers[cell.value.strip().lower()] = col_idx
            
            print(f"Znalezione nagłówki: {list(headers.keys())}")
            
            # Szukaj wymaganych kolumn
            required_keys = ['dlg_szt', 'ilosc', 'nazwa']
            missing_keys = [key for key in required_keys if key not in headers]
            if missing_keys:
                raise ValueError(f"Plik musi zawierać kolumny: {', '.join(missing_keys)}")
            
            # Kolumna Color (jeśli istnieje)
            color_col_idx = headers.get('color', None)
            
            # Przeszukaj wiersze z danymi
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False), 2):
                try:
                    # Pobierz wartości z właściwych kolumn
                    nazwa_cell = row[headers['nazwa'] - 1]
                    dlg_szt_cell = row[headers['dlg_szt'] - 1]
                    ilosc_cell = row[headers['ilosc'] - 1]
                    color_cell = row[color_col_idx - 1] if color_col_idx else None
                    
                    nazwa = str(nazwa_cell.value).strip() if nazwa_cell.value else None
                    color = str(color_cell.value).strip() if color_cell and color_cell.value else ""
                    
                    if not nazwa:
                        continue
                    
                    # Sprawdź czy zawiera "RAL" w kolumnie Color
                    if 'ral' not in color.lower():
                        continue
                    
                    # Parsuj wartości
                    dlg_szt = float(dlg_szt_cell.value or 0)
                    ilosc = int(ilosc_cell.value or 0)
                    
                    if dlg_szt > 0 and ilosc > 0:
                        profil = Profil(
                            nazwa=nazwa,
                            dlg_szt=dlg_szt,
                            ilosc=ilosc,
                            color=color
                        )
                        self.profile.append(profil)
                        print(f"✅ Załadowano: {nazwa} | DLG: {dlg_szt} | ILOSC: {ilosc} | COLOR: {color}")
                
                except (ValueError, TypeError, IndexError) as e:
                    print(f"⚠️ Wiersz {row_idx} - błąd parsowania: {str(e)}")
                    continue
            
            # Grupuj profile o tej samej nazwie
            self._group_profiles()
            
            print(f"\n✅ Załadowano {len(self.profile)} profili z oznaczeniem RAL")
            return len(self.profile) > 0
        
        except Exception as e:
            raise Exception(f"Błąd podczas wczytywania pliku: {str(e)}")
    
    def _group_profiles(self):
        """Grupuj profile o tej samej nazwie, sumując ich parametry"""
        grouped: Dict[str, Profil] = {}
        
        for profil in self.profile:
            key = profil.nazwa
            if key in grouped:
                # Sumuj metraż i ilość
                grouped[key].dlg_szt += profil.dlg_szt
                grouped[key].ilosc += profil.ilosc
            else:
                grouped[key] = profil
        
        self.profile = list(grouped.values())
        print(f"Po zgrupowaniu: {len(self.profile)} unikalnych profili")
    
    def calculate_total(self, kwota_okleinowania: float, dwustronne: bool = False, 
                       wspolczynnik: float = 1.0) -> Tuple[float, Dict[str, float]]:
        """Oblicz całkowitą cenę
        
        Args:
            kwota_okleinowania: cena za metr bieżący
            dwustronne: czy okleinowanie dwustronne
            wspolczynnik: współczynnik mnożenia dla okleinowania dwustronnego
        
        Returns:
            Tuple[cena_całkowita, dict_cen_profili]
        """
        ceny_profili = {}
        total = 0.0
        
        for profil in self.profile:
            cena = profil.calculate_price(kwota_okleinowania, dwustronne, wspolczynnik)
            if cena > 0:
                ceny_profili[profil.nazwa] = cena
                total += cena
        
        return total, ceny_profili
    
    def get_profiles(self) -> List[Profil]:
        """Zwróć listę profili"""
        return self.profile
    
    def toggle_all(self, state: bool):
        """Zaznacz/odznacz wszystkie profile"""
        for profil in self.profile:
            profil.wybrany = state
